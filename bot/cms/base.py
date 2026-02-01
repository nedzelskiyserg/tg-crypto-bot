"""
Базовый модуль для работы с CMS (Content Management System)
Парсит XLSX файлы и формирует динамическое меню для бота
"""
import os
import logging
from typing import Dict, List, Optional
from dataclasses import dataclass
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class MenuItem:
    """Класс для представления пункта меню"""
    id: str
    name: str
    parent_id: Optional[str]
    menu_type: str  # 'mess_button', 'message', 'reply_button' или 'reply_button_line'
    text: str
    callback_data: str
    order: int = 0
    group: Optional[str] = None  # Группа для reply_button_line (для группировки в ряды)
    
    def __post_init__(self):
        """Валидация данных"""
        if not self.callback_data:
            self.callback_data = f"menu_{self.id}"


class CMS:
    """Класс для управления контентом из XLSX файла"""
    
    def __init__(self, xlsx_path: str = "menu.xlsx"):
        """
        Инициализация CMS
        
        Args:
            xlsx_path: Путь к XLSX файлу с меню
        """
        self.xlsx_path = xlsx_path
        self.menu_items: Dict[str, MenuItem] = {}
        self.load_menu()
    
    def load_menu(self) -> None:
        """Загружает меню из XLSX файла"""
        if not os.path.exists(self.xlsx_path):
            logger.warning(f"Файл меню {self.xlsx_path} не найден. Создайте его по примеру.")
            return
        
        try:
            workbook = load_workbook(self.xlsx_path)
            sheet = workbook.active
            
            # Пропускаем заголовок (первая строка)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1]:  # Пропускаем пустые строки
                    continue
                
                try:
                    # Обработка типа меню (поддержка старого 'button' для обратной совместимости)
                    menu_type_raw = str(row[3]).strip().lower() if row[3] else 'mess_button'
                    menu_type = 'mess_button'  # По умолчанию
                    if menu_type_raw in ['mess_button', 'message', 'reply_button', 'reply_button_line', 'button']:
                        if menu_type_raw == 'button':
                            menu_type = 'mess_button'  # Конвертируем старый тип
                        else:
                            menu_type = menu_type_raw
                    
                    # Группа (для reply_button_line) - столбец 7 (если есть)
                    group = None
                    if len(row) > 7 and row[7]:
                        group = str(row[7]).strip() if row[7] else None
                    
                    item = MenuItem(
                        id=str(row[0]).strip(),
                        name=str(row[1]).strip(),
                        parent_id=str(row[2]).strip() if row[2] else None,
                        menu_type=menu_type,
                        text=str(row[4]).strip() if row[4] else '',
                        callback_data=str(row[5]).strip() if row[5] else f"menu_{row[0]}",
                        order=int(row[6]) if row[6] and str(row[6]).isdigit() else 0,
                        group=group
                    )
                    
                    self.menu_items[item.id] = item
                    
                except Exception as e:
                    logger.error(f"Ошибка при обработке строки {row}: {e}")
                    continue
            
            workbook.close()
            logger.info(f"Загружено {len(self.menu_items)} пунктов меню из {self.xlsx_path}")
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке меню из {self.xlsx_path}: {e}")
            raise
    
    def reload_menu(self) -> None:
        """Перезагружает меню из файла"""
        self.menu_items.clear()
        self.load_menu()
    
    def get_menu_item(self, item_id: str) -> Optional[MenuItem]:
        """Получает пункт меню по ID"""
        return self.menu_items.get(item_id)
    
    def get_children(self, parent_id: Optional[str] = None) -> List[MenuItem]:
        """
        Получает дочерние элементы меню
        
        Args:
            parent_id: ID родительского элемента (None для корневых элементов)
        
        Returns:
            Список дочерних элементов, отсортированных по order
        """
        children = [
            item for item in self.menu_items.values()
            if item.parent_id == parent_id
        ]
        return sorted(children, key=lambda x: x.order)
    
    def get_root_items(self) -> List[MenuItem]:
        """Получает корневые элементы меню (без родителя)"""
        return self.get_children(None)
    
    def get_menu_path(self, item_id: str) -> List[MenuItem]:
        """
        Получает путь к элементу меню (от корня до элемента)
        
        Args:
            item_id: ID элемента
        
        Returns:
            Список элементов от корня до указанного элемента
        """
        path = []
        current_id = item_id
        
        while current_id:
            item = self.get_menu_item(current_id)
            if not item:
                break
            path.insert(0, item)
            current_id = item.parent_id
        
        return path
    
    def get_all_items(self) -> List[MenuItem]:
        """Получает все элементы меню"""
        return list(self.menu_items.values())
    
    def find_item_by_name_and_type(self, name: str, menu_type: str) -> Optional[MenuItem]:
        """
        Находит элемент меню по названию и типу
        
        Args:
            name: Название элемента
            menu_type: Тип элемента
        
        Returns:
            MenuItem или None
        """
        for item in self.menu_items.values():
            if item.name == name and item.menu_type == menu_type:
                return item
        return None
